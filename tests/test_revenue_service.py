import pytest
from datetime import date
from openpyxl import load_workbook
import io

from app.services.user_service import UserService
from app.services.store_service import StoreService
from app.services.revenue_service import RevenueService
from app.models.revenue import Revenue


@pytest.mark.asyncio
async def test_revenue_service(session):

    user_svc = UserService(session)
    store_svc = StoreService(session)
    rev_svc = RevenueService(session)

    user = await user_svc.get_or_create("Jane", "Smith", "manager")
    store = await store_svc.get_or_create("Store1")

    today = date.today()
    revenue = await rev_svc.create_revenue(150.0, store.id, user.id, today)
    assert isinstance(revenue, Revenue)
    assert revenue.amount == 150.0
    assert revenue.store_id == store.id
    assert revenue.manager_id == user.id

    excel_bytes, images = await rev_svc.export_report()
    assert isinstance(excel_bytes, (bytes, bytearray))

    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    sheets = wb.sheetnames
    assert "Details" in sheets
    assert "Summary" in sheets

    assert isinstance(images, dict)
    assert "Store1" in images
    assert isinstance(images["Store1"], (bytes, bytearray))
